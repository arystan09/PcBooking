import json
import math
import random
import re
from datetime import time
from django.views.generic import ListView, CreateView, FormView, DetailView
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import redirect, render
from django.urls import reverse
from django.core.paginator import Paginator
import requests
from django.conf import settings
from selenium import webdriver
from .models import Club, User
from .forms import *
from django.contrib.auth import authenticate,login,logout
from django.contrib.admin import AdminSite
from django.contrib import admin 
from django.utils.text import capfirst
from django.contrib import messages


def sidebar(request):
    admin_site = AdminSite()
    available_apps = []

    app_list = admin.site.get_app_list(request)
    for app in app_list:
        app_name = app['name']
        app_label = app['app_label']
        app_url = reverse('admin:index') + f"{app_label}/"
        models_list = []
        for model_dict in app['models']:
            model = model_dict.get('model')  # Get the model class if it exists
            model_admin = admin_site._registry.get(model)
            if model:
                app_label = model._meta.app_label
                info = (app_label, model._meta.model_name)
                model_info = {
                    "model": model,
                    "name": capfirst(model._meta.verbose_name_plural),
                    "object_name": model._meta.object_name,
                    # "perms": perms,
                    "admin_url": None,
                    "add_url": None,
                }
                model_info["admin_url"] = reverse(
                            "admin:%s_%s_changelist" % info, current_app=capfirst(model._meta.verbose_name_plural)
                        )
                model_info["add_url"] = reverse(
                            "admin:%s_%s_add" % info, current_app=capfirst(model._meta.verbose_name_plural)
                        )

                models_list.append(model_info)
        available_apps.append(
            {
                'name': app_name,
                'models': models_list,
                'app_label': app_label,
                'app_url': app_url
            }
        )

    data = {
        "has_permission": True,
        "is_popup": False,
        "available_apps": available_apps,
        "site_title": admin_site.site_title,
        "site_header": admin_site.site_header,
    }

    admin_context = admin_site.each_context(request)
    admin_context.update(data)
    return admin_context

#example of full address
#
def geocode_address(address):
    # api_key = settings.OPENCAGE_API_KEY
    api_key = ""
    url = f"https://api.opencagedata.com/geocode/v1/json?q={address}&key={api_key}"
    
    response = requests.get(url)
    data = response.json()

    if data['results']:
        latitude = data['results'][0]['geometry']['lat']
        longitude = data['results'][0]['geometry']['lng']
        return latitude, longitude
    else:
        return None, None
    
def haversine(lat1, lon1, lat2, lon2):
    """
    Calculate the great-circle distance between two points on the Earth surface.
    Returns distance in kilometers.
    """
    R = 6371  # Earth radius in kilometers

    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)

    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2.0)**2 + \
        math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0)**2

    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    meters = R * c
    return meters

from haversine import haversine, Unit

def list_computer_clubs(request):
    clubs = Club.objects.all()
    latitude = None
    longitude = None
    radius = None
    if request.method == "POST":
        if "latitude" and "longitude" in request.POST:
            data = json.loads(request.body)
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            radius = request.GET.get('radius')
            print(latitude, longitude)
    search_form = SearchForm(request.GET or None)
    if search_form.is_valid():
        query = search_form.cleaned_data['query']
        if query:
            clubs = clubs.filter(Q (name__icontains=query) | Q(address__icontains=query))

    distance = {}
    if latitude is not None and longitude is not None:
        user = request.user
        user.last_latitude = latitude
        user.last_longitude = longitude
        user.save()
        # user_location = Point(float(longitude), float(latitude), srid=4326)
        
        
        print(f"res: {latitude}, {longitude}")
            
    paginator = Paginator(clubs, 8)  # Show 10 clubs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
    }
    return render(request, 'base/allcompclubs.html', context)

def detailed_club_view(request,club_id):
    club = get_object_or_404(Club, id=club_id)
    gis_api_key = settings.GIS_API_KEY
    context = {
        'club': club,
        "gis_api_key": gis_api_key,
    }

    return render(request, 'base/clubdetailed.html', context)

# def dashboard(request):
#     user_profile = User.objects.get(user=request.user)
#     # upcoming_bookings = Booking.objects.filter(user=request.user, date__gte=timezone.now()).order_by('date')
#     context = {
#         'user_profile': user_profile,
#         # 'upcoming_bookings': upcoming_bookings,
#     }
#     return render(request, 'dashboard.html', context)

def home(request):
    return render(request, 'base/home.html')


# def book(request):
#     return render(request, 'base/book.html')
    

def update_user_profile(request):
    if not request.user.is_authenticated:
        return redirect('login')
    current_user = User.objects.get(id=request.user.id)
    form = RegisterForm(request.POST or None,instance=current_user)
    if form.is_valid():
        form.save()
        return redirect('user-profile')
    
    return render(request, "base/update_user_profile.html",{ 'form': form})


def register(request):
    form = RegisterForm()
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    return render(request,'base/register.html', {'form': form})

def login_user(request):
    form = LoginUserForm()

    if request.method == 'POST':
        form = LoginUserForm(request,data=request.POST)
        if form.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request,username=username, password=password)

            if user is not None:
                login(request, user)
                return redirect('home')

    return render(request, 'base/login.html', {'form': form})


def logout_user(request):
    logout(request)
    return redirect('login')

def user_profile(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'base/user_profile.html')

from openpyxl import load_workbook as lw

def upload_club_file(request):
    if request.method == "POST":
        uploaded_file = request.FILES.get('file', None)
        
        if not uploaded_file:
            messages.error(request, 'No file uploaded. Please upload a file.')
        else:
            if not uploaded_file.name.endswith('.xlsx'):
                messages.error(request, 'Please upload an .xlsx file.')
            else:
                load_workbook = lw(uploaded_file)
                clubs = load_workbook.active
                messages.success(request, 'File uploaded and processed successfully.')
                data_length = 1
                for i in range(1,168):
                    try:
                        club_name = clubs["A"+str(i)].value
                        club_description = clubs["B"+str(i)].value
                        club_address = clubs["D"+str(i)].value
                        club_district = clubs["H" + str(i)].value
                        club_city = clubs["I"+str(i)].value
                        working_hours = clubs["M"+str(i)].value
                        time_range = re.search(r'(\d{2}:\d{2})-(\d{2}:\d{2})', working_hours)
                        if time_range:
                            start_time_str, end_time_str = time_range.groups()
                            # print("start_time",start_time_str)
                            # print("end_time",end_time_str)
                            start_time = time.fromisoformat(start_time_str)
                            if end_time_str == "24:00":
                                end_time = time(23, 59)
                            else:
                                end_time = time.fromisoformat(end_time_str)
                        club_work_time_start = start_time
                        club_work_time_end = end_time
                        club_rating = random.uniform(2, 5)
                        club_phone_number = clubs["Q" + str(i)].value
                        club_website = clubs["T"+str(i)].value
                        club_instagram_url = clubs["V"+str(i)].value
                        club_whatsapp_url = clubs["Z"+str(i)].value
                        club_longitude = clubs["AE"+str(i)].value
                        club_latitude = clubs["AD"+str(i)].value
                        club_2gis_url = clubs["AF"+str(i)].value
                        club_owner = request.user
                        club_x_size = 200
                        club_y_size = 200
                        if club_description is None:
                            club_description = ""
                        
                        Club.objects.create(
                            name=club_name,
                            description=club_description,
                            address=club_address,
                            city=club_city,
                            district = club_district,
                            work_time_start = club_work_time_start,
                            work_time_end = club_work_time_end,
                            x_size=club_x_size,
                            y_size=club_y_size,
                            longitude = club_longitude,
                            latitude = club_latitude,
                            phone_number=club_phone_number,
                            instagram_url=club_instagram_url,
                            whatsapp_url=club_whatsapp_url,
                            rating=club_rating,
                            website=club_website,
                            twogis_url=club_2gis_url,
                            user = club_owner
                        )
                    except Exception as E:
                        print("error in line: ",i, E)
    context = {}
    context.update(sidebar(request))
    return render(request, 'base/upload_club_file.html', context)
